
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD
from flask import Response
from reportlab.pdfgen import canvas
import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado, foto_empleado, salario_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['sexo_empleado'],
                           dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['profesion_empleado'], result_foto_perfil, salario_entero)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Telefono", "Email", "Profesión", "Salario", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     salario_empleado, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

#reporte Usuarios Excel
import os
import datetime
import openpyxl
from flask import send_file

import os
import datetime
import openpyxl
from flask import send_file

def generarReporteUsuariosExcel():
    dataUsuarios = obtenerUsuarios()  # Utiliza tu función para obtener los usuarios
    
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre y Apellido", "Correo Electrónico", "Fecha de Creación")
    hoja.append(cabeceraExcel)

    # Formato para fechas
    formato_fecha = 'YYYY-MM-DD'

    # Agregar los registros de los usuarios a la hoja
    for registro in dataUsuarios:
        nombre_usuario = registro['name_surname']
        email_usuario = registro['email_user']
        fecha_creacion = registro['created_user']

        # Si la fecha de creación no es un objeto datetime, convertirlo
        if isinstance(fecha_creacion, str):
            fecha_creacion = datetime.datetime.strptime(fecha_creacion, "%Y-%m-%d %H:%M:%S")
        
        hoja.append((nombre_usuario, email_usuario, fecha_creacion))

        # Formato de fecha en la columna de "Fecha de Creación"
        for fila_num in range(2, hoja.max_row + 1):
            columna = 3  # Columna C (Fecha de Creación)
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_fecha  # Establecer el formato de fecha

    # Ajuste automático del tamaño de las columnas
    for col in hoja.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Añadir un pequeño margen
        hoja.column_dimensions[column].width = adjusted_width

    # Nombre del archivo Excel con la fecha actual
    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_usuarios_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    
    # Ruta donde se guardará el archivo
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(os.path.abspath(__file__)), carpeta_descarga)

    # Crear la carpeta si no existe
    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        os.chmod(ruta_descarga, 0o755)

    # Ruta completa para guardar el archivo Excel
    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo Excel como respuesta
    return send_file(ruta_archivo, as_attachment=True)

#generar PDF
def generarReportePDF():
    empleados = empleadosReporte()

    if not empleados:
        return "No hay empleados para generar el PDF."

    # Configurar la respuesta para la descarga del archivo
    response = Response(content_type='application/pdf')
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_empleados.pdf'

    # Crear el PDF
    pdf = canvas.Canvas(response.stream)
    pdf.setTitle("Reporte de Empleados")

    # Encabezado
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(200, 800, "Reporte de Empleados")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(200, 785, f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    # Columnas
    columnas = ["ID", "Nombre", "Apellido", "Sexo", "Salario", "Profesión"]
    
    # Ancho de las columnas
    x_inicial = 50
    y_inicial = 750
    espacio_columna = [50, 100, 100, 70, 90, 130]  # Columnas de ancho ajustado

    # Dibujar encabezados
    pdf.setFont("Helvetica-Bold", 10)
    for i, columna in enumerate(columnas):
        pdf.drawString(x_inicial + sum(espacio_columna[:i]), y_inicial, columna)

    # Dibujar datos
    pdf.setFont("Helvetica", 9)
    y_actual = y_inicial - 20
    for empleado in empleados:
        datos = [
            str(empleado["id_empleado"]),
            empleado["nombre_empleado"],
            empleado["apellido_empleado"],
            empleado["sexo_empleado"],
            f"${empleado['salario_empleado']:,.0f}",  # Formato moneda
            empleado["profesion_empleado"]
        ]
        for i, dato in enumerate(datos):
            pdf.drawString(x_inicial + sum(espacio_columna[:i]), y_actual, dato)
        y_actual -= 20  # Espacio entre filas

        # Si llegamos al final de la página, crear nueva página
        if y_actual < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 9)
            y_actual = 750

    pdf.save()
    return response

#generar Pdf usuario 

def obtenerUsuarios():
    try:
        conexion = connectionBD()  # Usar tu función de conexión
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT id, name_surname, email_user, created_user FROM users")
        usuarios = cursor.fetchall()
    except Exception as e:
        print(f"Error al obtener usuarios: {e}")
        usuarios = []
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()
    
    return usuarios
def generarReporteUsuariosPDF():
    usuarios = obtenerUsuarios()

    if not usuarios:
        return "No hay usuarios para generar el PDF."

    response = Response(content_type='application/pdf')
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_usuarios.pdf'
    
    pdf = canvas.Canvas(response.stream)
    pdf.setTitle("Reporte de Usuarios")
    
    # Encabezado
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(200, 800, "Reporte de Usuarios")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(200, 785, f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    # Columnas
    columnas = ["Nombre y Apellido", "Email", "Fecha de Creación"]
    x_inicial = 50
    y_inicial = 750
    espacio_columna = [150, 150, 150]  
    
    # Dibujar encabezados
    pdf.setFont("Helvetica-Bold", 10)
    for i, columna in enumerate(columnas):
        pdf.drawString(x_inicial + sum(espacio_columna[:i]), y_inicial, columna)
    
    # Dibujar datos
    pdf.setFont("Helvetica", 9)
    y_actual = y_inicial - 20
    for usuario in usuarios:
        datos = [
            
            usuario["name_surname"],
            usuario["email_user"],
            usuario["created_user"].strftime('%d/%m/%Y %H:%M:%S')
        ]
        for i, dato in enumerate(datos):
            pdf.drawString(x_inicial + sum(espacio_columna[:i]), y_actual, dato)
        y_actual -= 20

        # Control de salto de página
        if y_actual < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 9)
            y_actual = 750
    
    pdf.save()
    return response


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.salario_empleado,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Extraer y procesar datos del formulario
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                sexo_empleado = data.form['sexo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                profesion_empleado = data.form['profesion_empleado']

                # Procesar salario eliminando caracteres no numéricos
                salario_sin_puntos = re.sub('[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                # Construir consulta SQL y parámetros dinámicamente
                query_base = """
                    UPDATE tbl_empleados
                    SET 
                        nombre_empleado = %s,
                        apellido_empleado = %s,
                        sexo_empleado = %s,
                        telefono_empleado = %s,
                        email_empleado = %s,
                        profesion_empleado = %s,
                        salario_empleado = %s
                """
                params = [
                    nombre_empleado, apellido_empleado, sexo_empleado,
                    telefono_empleado, email_empleado, profesion_empleado, salario_empleado
                ]

                # Verificar si se subió un archivo de foto
                if 'foto_empleado' in data.files and data.files['foto_empleado'].filename != '':
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)
                    query_base += ", foto_empleado = %s"
                    params.append(fotoForm)

                # Agregar condición WHERE
                query_base += " WHERE id_empleado = %s"
                params.append(id_empleado)

                # Ejecutar la consulta
                cursor.execute(query_base, params)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
